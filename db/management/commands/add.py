import datetime
from time import sleep
from django.core.management.base import BaseCommand
from db.models import Product
from utils.api_and_crawl import Biid , Mobo
from utils.convert import mobo_to_biid , hash_product
import json
import pandas as pd
import openpyxl


class Command(BaseCommand):

    @staticmethod
    def product_exists(mobo_product):
        try:
            Product.objects.get(identifier=mobo_product)
            return True
        except:
            return False

    def handle(self ,*args ,**options):
        mo =Mobo()
        b =Biid()

        for mobo_product_short in mo.get_products():

            mobo_product_info = mo.get_info(mobo_product_short)
            for item in mobo_product_info:
                if item != "category" and item != "product_id":
                    print(int(item["stock"]))
                #     product_id = mobo_product_info["product_id"]
                #     product_identifier = f"{product_id}---{item}"
                #
                #     if self.product_exists(product_identifier):
                #         print(f"product id={product_id} already exists")
                #         continue
                #
                #     category = [mobo_product_info["category"]]
                #     mobo_product = mobo_product_info[item]
                #
                #     if category[0] == "قاب موبایل":
                #         if mobo_product["status_title"] == "yes":
                #             category = [mobo_product["category_var"] ,mobo_product["category_name"]]
                #
                #     biid_product, colors = mobo_to_biid(mobo_product, category)
                #
                #     if not biid_product['stock']:
                #         print(f"product id={product_id} is out of stock")
                #         continue
                #
                #     images_url = mo.get_images(product_id)
                #     if not images_url:
                #         print(f"product id={product_id} has no image")
                #         continue
                #
                #     if len(colors) == 0:
                #         biid_product["has_variants"] = False
                #
                #     product_id = b.add_product(biid_product)
                #
                #     p = Product.objects.create(id=product_id,
                #                                identifier=product_identifier,
                #                                from_mobo=True,
                #                                )
                #
                #     for image_url in images_url:
                #         try:
                #             b.add_image_to_product(product_id, image_url=image_url, image_alt=biid_product['name'])
                #         except:
                #             pass
                #
                #     if len(colors) != 0:
                #         b.add_colors_to_product(product_id, colors)
                #         variants = b.get_product_variants(product_id)
                #         for variant in variants:
                #             b.update_product_variant(product_id, variant['id'],
                #                                      {'product_identifier': biid_product['barcode']})
                #
                #     try:
                #         if mobo_product_info[item]["status_title"] == "no":
                #             file_path = "productId.xlsx"
                #             wb = openpyxl.load_workbook(file_path)
                #             ws = wb.active
                #             last_row = ws.max_row + 1
                #             ws.cell(row=last_row, column=1).value = product_id
                #             wb.save(file_path)
                #     except:
                #         pass
                #
                #     value = biid_product['name']
                #     pro_id = product_id
                #     b.add_tag(product_id=pro_id, value=value)
                #     p.commit = True
                #     p.save()
                #     print(
                #         f"mobo product id={mobo_product_info['product_id']} successfully added to biid product id={product_id}")




















